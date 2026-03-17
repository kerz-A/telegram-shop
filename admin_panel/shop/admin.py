import logging
from io import BytesIO

from django.contrib import admin, messages
from django.db.models import Count, Sum
from django.http import HttpResponse

import httpx
import openpyxl
from openpyxl.styles import Font, Alignment

from shop.models import (
    BotSettings,
    Broadcast,
    CartItem,
    Category,
    Customer,
    FAQ,
    Order,
    OrderItem,
    Product,
    ProductImage,
)
from shop_core.enums import BroadcastStatus, OrderStatus
from shop_core.schemas import BroadcastRequest, OrderNotification, WebhookPayload
from shop_core.constants import ACTION_ORDER_STATUS_CHANGED, ACTION_START_BROADCAST

from django.conf import settings

logger = logging.getLogger(__name__)


def _notify_bot(payload: WebhookPayload) -> bool:
    """Send webhook to bot service."""
    try:
        resp = httpx.post(
            settings.BOT_WEBHOOK_URL,
            json=payload.model_dump(),
            timeout=5.0,
        )
        return resp.status_code == 200
    except Exception as e:
        logger.error("Failed to notify bot: %s", e)
        return False


# ─── Inlines ────────────────────────────────────────────────────────


class ProductImageInline(admin.TabularInline):
    model = ProductImage
    extra = 1
    fields = ["image", "sort_order"]


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ["product", "product_name", "price", "quantity", "total"]
    can_delete = False

    def total(self, obj):
        val = obj.total
        return f"{val}₽" if val is not None else "—"

    total.short_description = "Сумма"


class OrderInlineForCustomer(admin.TabularInline):
    model = Order
    extra = 0
    fields = ["id", "status", "total", "created_at"]
    readonly_fields = ["id", "status", "total", "created_at"]
    can_delete = False
    show_change_link = True
    verbose_name = "Заказ"
    verbose_name_plural = "История заказов"


# ─── Category ───────────────────────────────────────────────────────


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ["name", "parent", "sort_order", "is_active", "product_count"]
    list_filter = ["is_active", "parent"]
    search_fields = ["name", "slug"]
    prepopulated_fields = {"slug": ("name",)}
    list_editable = ["sort_order", "is_active"]

    def product_count(self, obj):
        return obj.products.count()

    product_count.short_description = "Товаров"


# ─── Product ────────────────────────────────────────────────────────


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ["name", "category", "price", "is_active", "image_count", "created_at"]
    list_filter = ["is_active", "category"]
    search_fields = ["name", "description"]
    list_editable = ["price", "is_active"]
    inlines = [ProductImageInline]

    def image_count(self, obj):
        return obj.images.count()

    image_count.short_description = "Фото"


# ─── Customer ───────────────────────────────────────────────────────


@admin.register(Customer)
class CustomerAdmin(admin.ModelAdmin):
    list_display = [
        "telegram_id",
        "username",
        "first_name",
        "phone",
        "is_admin",
        "order_count",
        "total_spent",
        "created_at",
    ]
    list_filter = ["is_admin"]
    search_fields = ["telegram_id", "username", "first_name", "phone"]
    readonly_fields = ["telegram_id", "username", "first_name", "last_name", "phone", "created_at"]
    inlines = [OrderInlineForCustomer]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.annotate(
            _order_count=Count("orders"),
            _total_spent=Sum("orders__total"),
        )

    def order_count(self, obj):
        return obj._order_count

    order_count.short_description = "Заказов"
    order_count.admin_order_field = "_order_count"

    def total_spent(self, obj):
        val = obj._total_spent
        return f"{val}₽" if val else "0₽"

    total_spent.short_description = "Сумма покупок"
    total_spent.admin_order_field = "_total_spent"


# ─── Order ──────────────────────────────────────────────────────────


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ["id", "customer", "status", "total", "created_at"]
    list_filter = ["status"]
    search_fields = ["id", "customer__telegram_id", "customer__username", "full_name"]
    readonly_fields = ["customer", "full_name", "phone", "address", "total", "created_at"]
    inlines = [OrderItemInline]
    actions = ["export_paid_orders"]

    def save_model(self, request, obj, form, change):
        """Notify bot when status changes."""
        if change:
            old_obj = Order.objects.get(pk=obj.pk)
            status_changed = old_obj.status != obj.status
        else:
            status_changed = False

        super().save_model(request, obj, form, change)

        if status_changed:
            notification = OrderNotification(
                order_id=obj.pk,
                customer_telegram_id=obj.customer.telegram_id,
                new_status=OrderStatus(obj.status),
            )
            payload = WebhookPayload(
                action=ACTION_ORDER_STATUS_CHANGED,
                data=notification.model_dump(),
            )
            success = _notify_bot(payload)
            if success:
                messages.success(request, f"Уведомление отправлено клиенту о статусе: {obj.get_status_display()}")
            else:
                messages.warning(request, "Не удалось отправить уведомление клиенту")

    @admin.action(description="📥 Экспорт оплаченных заказов в Excel")
    def export_paid_orders(self, request, queryset):
        paid_orders = queryset.filter(status=OrderStatus.PAID)
        if not paid_orders.exists():
            messages.warning(request, "Нет оплаченных заказов для экспорта")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Оплаченные заказы"

        headers = ["ID", "Клиент", "ФИО", "Телефон", "Адрес", "Сумма", "Дата"]
        header_font = Font(bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, order in enumerate(paid_orders.select_related("customer"), 2):
            ws.cell(row=row_idx, column=1, value=order.pk)
            ws.cell(row=row_idx, column=2, value=str(order.customer))
            ws.cell(row=row_idx, column=3, value=order.full_name)
            ws.cell(row=row_idx, column=4, value=order.phone)
            ws.cell(row=row_idx, column=5, value=order.address)
            ws.cell(row=row_idx, column=6, value=float(order.total))
            ws.cell(row=row_idx, column=7, value=order.created_at.strftime("%Y-%m-%d %H:%M"))

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="paid_orders.xlsx"'
        return response


# ─── FAQ ────────────────────────────────────────────────────────────


@admin.register(FAQ)
class FAQAdmin(admin.ModelAdmin):
    list_display = ["question", "is_active", "popularity_score"]
    list_filter = ["is_active"]
    search_fields = ["question", "answer"]
    list_editable = ["is_active", "popularity_score"]


# ─── BotSettings ────────────────────────────────────────────────────


@admin.register(BotSettings)
class BotSettingsAdmin(admin.ModelAdmin):
    list_display = ["__str__", "admin_chat_id", "updated_at"]

    def has_add_permission(self, request):
        # Singleton: allow add only if none exists
        return not BotSettings.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False


# ─── Broadcast ──────────────────────────────────────────────────────


@admin.register(Broadcast)
class BroadcastAdmin(admin.ModelAdmin):
    list_display = ["id", "text_preview", "status", "sent_count", "error_count", "created_at"]
    list_filter = ["status"]
    readonly_fields = ["sent_count", "error_count"]
    actions = ["send_broadcast"]

    def text_preview(self, obj):
        return obj.text[:60] + "..." if len(obj.text) > 60 else obj.text

    text_preview.short_description = "Текст"

    @admin.action(description="📤 Отправить рассылку")
    def send_broadcast(self, request, queryset):
        for broadcast in queryset:
            if broadcast.status not in (BroadcastStatus.DRAFT, BroadcastStatus.READY):
                messages.warning(request, f"Рассылка #{broadcast.pk} уже отправлена или отправляется")
                continue

            broadcast.status = BroadcastStatus.READY
            broadcast.save(update_fields=["status"])

            br_data = BroadcastRequest(
                broadcast_id=broadcast.pk,
                text=broadcast.text,
                image_path=broadcast.image.path if broadcast.image else None,
            )
            payload = WebhookPayload(
                action=ACTION_START_BROADCAST,
                data=br_data.model_dump(),
            )
            success = _notify_bot(payload)
            if success:
                messages.success(request, f"Рассылка #{broadcast.pk} запущена")
            else:
                messages.error(request, f"Не удалось запустить рассылку #{broadcast.pk}")


# ─── Admin site config ──────────────────────────────────────────────

admin.site.site_header = "🛒 Shop Admin"
admin.site.site_title = "Shop Admin"
admin.site.index_title = "Управление магазином"
